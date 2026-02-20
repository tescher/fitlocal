import io
from openpyxl import Workbook
from models import WorkoutSession, LoggedSet, PlannedWorkout


def generate_xlsx(user_id):
    wb = Workbook()
    ws = wb.active
    ws.title = "Workout Log"

    headers = ["Date", "Workout Name", "Exercise", "Set", "Weight (lbs)", "Reps", "RPE", "Notes"]
    ws.append(headers)

    for col in range(1, len(headers) + 1):
        ws.cell(row=1, column=col).font = ws.cell(row=1, column=col).font.copy(bold=True)

    sessions = (
        WorkoutSession.query
        .filter_by(user_id=user_id)
        .order_by(WorkoutSession.date.asc())
        .all()
    )

    for session in sessions:
        workout_name = ""
        if session.planned_workout:
            workout_name = session.planned_workout.workout_name

        for logged_set in sorted(session.logged_sets, key=lambda s: (s.exercise_name, s.set_number)):
            ws.append([
                session.date.strftime("%Y-%m-%d") if session.date else "",
                workout_name,
                logged_set.exercise_name,
                logged_set.set_number,
                logged_set.weight_lbs,
                logged_set.reps_completed,
                logged_set.rpe,
                logged_set.notes or "",
            ])

    # Auto-size columns
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_length + 2, 40)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output
